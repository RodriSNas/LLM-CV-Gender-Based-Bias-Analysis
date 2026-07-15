import re
import pandas as pd
from scipy.stats import spearmanr
from openpyxl.styles import PatternFill


HUMAN_SCORES_CSV   = "datasets/Resume/validation_scores_rodrigo.csv" #path of the excel file with the human scores given
LLM_CSV       = "results/openai/scores_summary/scores_summary_openai_sem_fairness_inst.csv"
LLM_CSV_2     = "results/claude/scores_summary/scores_summary_claude_sem_fairness_inst.csv"
OUTPUT_XLSX   = "results/human_vs_LLM_scores_analysis.xlsx"

TECHNIQUES = [
    "zero_shot", "few_shot", "CoT", "ThoT",
    "least_to_most", "take_a_step_back", "self_consistency",
]

LLM_VERSION = "original"
HIGHLIGHT_TECHNIQUE = "zero_shot"
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def load_human_scores(filepath: str) -> pd.DataFrame:
    """
    Extracts (row_num, cv_id, domain, human_score) from the messy CSV
    using a regex anchored on the start of each record, since the
    Reasoning column breaks naive comma/csv parsing.
    """
    with open(filepath, encoding="latin1") as f:
        text = f.read()

    pattern = re.compile(
        r'^"?(?P<num>\d+),(?P<id>\d+),(?P<cat>[A-Z\-]+),(?P<score>\d+),',
        re.MULTILINE,
    )
    matches = pattern.findall(text)

    df = pd.DataFrame(matches, columns=["row_num", "cv_id", "domain", "human_score"])
    df["cv_id"] = df["cv_id"].astype(str)
    df["human_score"] = df["human_score"].astype(float)
    df["row_num"] = df["row_num"].astype(int)

    print(f"Parsed {len(df)} human-scored CVs from {filepath}.")
    return df


def load_llm_scores(filepath: str, version: str, model_name: str) -> pd.DataFrame:
    df = pd.read_csv(filepath, dtype={"cv_id": str})
    df = df[(df["version"] == version) & (df["success"] == True)].copy()

    df = df[
        ((df["technique"] != "self_consistency") & (df["sample_index"] == 0))
        | ((df["technique"] == "self_consistency") & (df["sample_index"] == -1))
    ]

    pivot = df.pivot_table(
        index="cv_id", columns="technique", values="score", aggfunc="first"
    ).reset_index()

    ordered_cols = ["cv_id"] + [t for t in TECHNIQUES if t in pivot.columns]
    pivot = pivot[ordered_cols]

    pivot.columns = ["cv_id"] + [f"llm_{t}_{model_name}" for t in ordered_cols[1:]]

    return pivot


def merge_scores(human_df: pd.DataFrame, llm_df: pd.DataFrame) -> pd.DataFrame:
    merged = human_df.merge(llm_df, on="cv_id", how="left")

    missing = merged[merged.filter(like="llm_").isna().all(axis=1)]
    if len(missing) > 0:
        print(f"\nWARNING: {len(missing)} CV(s) had no matching LLM scores:")
        print(missing[["cv_id", "domain"]].to_string(index=False))

    merged = merged.sort_values(["domain", "row_num"]).reset_index(drop=True)
    return merged


def compute_summary_stats(merged: pd.DataFrame, model_names: list) -> pd.DataFrame:
    rows = []
    for model_name in model_names:
        for technique in TECHNIQUES:
            col = f"llm_{technique}_{model_name}"
            if col not in merged.columns:
                continue

            pair = merged[["human_score", col]].dropna()
            n = len(pair)
            if n < 2:
                mae, rho, pval = float("nan"), float("nan"), float("nan")
            else:
                mae = (pair["human_score"] - pair[col]).abs().mean()
                rho, pval = spearmanr(pair["human_score"], pair[col])

            rows.append({
                "model": model_name,
                "technique": technique,
                "n": n,
                "MAE": round(mae, 4) if pd.notna(mae) else mae,
                "spearman_rho": round(rho, 4) if pd.notna(rho) else rho,
                "spearman_p": round(pval, 4) if pd.notna(pval) else pval,
            })

    summary = pd.DataFrame(rows)
    summary["technique"] = pd.Categorical(summary["technique"], categories=TECHNIQUES, ordered=True)
    summary = summary.sort_values(["model", "technique"]).reset_index(drop=True)
    return summary


def export_comparison(summary_df: pd.DataFrame, raw_df: pd.DataFrame, filepath: str) -> None:
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary (MAE & Spearman)", index=False)
        raw_df.to_excel(writer, sheet_name="Human vs LLM (raw)", index=False)

        ws = writer.sheets["Summary (MAE & Spearman)"]
        technique_col_idx = summary_df.columns.get_loc("technique") + 1  # 1-indexed for openpyxl

        for row_idx, technique in enumerate(summary_df["technique"], start=2):
            if technique == HIGHLIGHT_TECHNIQUE:
                for col_idx in range(1, len(summary_df.columns) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = HIGHLIGHT_FILL

    print(f"\nExported to: {filepath}")


if __name__ == "__main__":
    print("=" * 60)
    print("HUMAN VS LLM SCORE COMPARISON")
    print("=" * 60)

    print("\n[1/4] Loading human validation scores...")
    human_df = load_human_scores(HUMAN_SCORES_CSV)

    print("\n[2/4] Loading and pivoting LLM scores...")
    llm_gpt = load_llm_scores(LLM_CSV, LLM_VERSION, "gpt")
    llm_claude = load_llm_scores(LLM_CSV_2, LLM_VERSION, "claude")
    llm_df = llm_gpt.merge(llm_claude, on="cv_id", how="outer")

    print("\n[3/4] Merging...")
    merged = merge_scores(human_df, llm_df)

    print("\n[4/4] Computing MAE and Spearman's correlation per technique/model, and exporting...")
    summary = compute_summary_stats(merged, model_names=["gpt", "claude"])
    print(summary.to_string(index=False))
    export_comparison(summary, merged, OUTPUT_XLSX)

    print("\nDone.")
