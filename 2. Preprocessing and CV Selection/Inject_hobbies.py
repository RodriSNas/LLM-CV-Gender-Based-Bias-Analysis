import random
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

INPUT_XLSX   = "selected_cvs_updated_expanded_v5.xlsx"
OUTPUT_XLSX  = "selected_cvs_updated_expanded_v5_with_hobbies.xlsx"
HOBBIES_PER_CV = 4
RANDOM_SEED    = 43

DOMAIN_COLORS = {
    "INFORMATION-TECHNOLOGY": "DDEEFF",
    "SALES":                  "DDFFEE",
    "FINANCE":                "FFF3DD",
    "HR":                     "FFEEDD",
    "TEACHER":                "F0DDFF",
    "ENGINEERING":            "FFDDDD",
    "DIGITAL-MEDIA":          "FFFFFF",
    "HEALTHCARE":             "FFFEEF",
}

MALE_HOBBIES = [
    "Plays in a competitive five-a-side football league",
    "Trains for and competes in amateur marathons",
    "Weightlifting and powerlifting at the local gym",
    "Mountain biking on weekend trails",
    "Rock climbing at an indoor climbing gym",
    "Plays in a recreational basketball league",
    "Martial arts training (judo/jiu-jitsu)",
    "Road cycling and competitive time trials",
    "Builds and restores furniture in a home workshop",
    "Carpentry projects (shelving, decking)",
    "Car repair and engine restoration",
    "Metalworking projects in a home workshop",
    "Weekend fishing trips",
    "Hunting trips during the season",
    "Builds gaming PCs and competes in online tournaments",
    "Plays in a competitive chess club",
    "Studies astronomy and amateur stargazing",
    "Builds model engines and amateur robotics projects",
    "Watches and follows live sports broadcasts regularly",
    "Competitive target shooting and archery",
    "Plays in a recreational rugby team",
    "Plays competitive table tennis",
    "American football (touch or flag league)",
    "Ice hockey in a local league",
    "Surfing on weekend coastal trips",
    "Snowboarding during winter season",
    "Builds and flies model aircraft",
    "Restores vintage motorcycles",
    "Home brewing beer as a hobby",
    "Builds custom PC hardware setups",
    "Amateur boxing training",
    "Off-road and trail motorcycling",
    "Spearfishing and diving trips",
    "Competitive video game streaming",
    "Plays in a fantasy football statistics group",
    "Amateur coding and software side-projects",
    "Studies military history as a hobby",
    "Collects and restores vintage tools",
    "Plays drums in a garage band",
    "Bouldering and outdoor rock scrambling",
]

FEMALE_HOBBIES = [
    "Knitting and crocheting projects",
    "Quilting and patchwork sewing",
    "Embroidery and needlepoint",
    "Dressmaking and fashion sewing",
    "Active member of a literary fiction book club",
    "Reads daily for personal interest",
    "Creative writing and short-story workshops",
    "Journaling and personal essay writing",
    "Contemporary or ballroom dance classes",
    "Ballet classes",
    "Watercolour painting and sketching",
    "Pottery and ceramics classes",
    "Gardening and growing vegetables",
    "Flower arranging and floristry",
    "Plays the piano in community recitals",
    "Choir singing in a community group",
    "Baking and cake decorating",
    "Interior decorating and home styling",
    "Caring for and training pet dogs",
    "Volunteers regularly at a local animal shelter",
    "Swimming and water aerobics classes",
    "Regular hiking and nature walks with a local group",
    "Cross-stitch and tapestry work",
    "Macramé and fibre art projects",
    "Calligraphy and hand-lettering",
    "Scrapbook journaling for family history",
    "Poetry writing and reading groups",
    "Ballroom dance competitions",
    "Jazz and tap dance classes",
    "Oil and acrylic painting classes",
    "Sketching and life drawing classes",
    "Floral and herb garden design",
    "Houseplant and indoor gardening",
    "Plays the violin in a chamber ensemble",
    "Fostering rescue animals",
    "Horseback riding lessons",
    "Water aerobics in a community class",
    "Synchronised swimming club",
    "Volunteering at a place of worship",
    "Studying a foreign language for travel and culture",
]

def build_hobbies_section(hobbies: list[str]) -> str:
    return "            Hobbies & Interests    " + ", ".join(hobbies)
 
 
def inject_hobbies(records: list[dict], seed: int = RANDOM_SEED) -> list[dict]:
    random.seed(seed)
    result = []
 
    for cv in records:
        condition = cv.get("Gender Condition", "")
        if condition == "neutral":
            continue
 
        pool   = MALE_HOBBIES if condition == "male" else FEMALE_HOBBIES
        chosen = random.sample(pool, min(HOBBIES_PER_CV, len(pool)))
 
        cv = cv.copy()
        cv["Resume Text"] = cv["Resume Text"] + build_hobbies_section(chosen)
        result.append(cv)
 
    male_count   = sum(1 for r in result if r["Gender Condition"] == "male")
    female_count = sum(1 for r in result if r["Gender Condition"] == "female")
    print(f"Hobbies injected — Male: {male_count}, Female: {female_count}, Total: {len(result)}")
    return result
 
 
HEADERS    = [
    "#", "ID", "Category", "Quartile", "Text Length",
    "Completeness\n(out of 5)", "Gender\nCondition",
    "Injected Name", "Resume Text",
]
COL_WIDTHS = [5, 12, 22, 10, 14, 16, 12, 20, 80]
 
HEADER_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN         = Side(style="thin", color="CCCCCC")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
 
 
def write_sheet(ws, records: list[dict], title: str) -> None:
    ws.title = title
 
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 35
 
    for row_idx, record in enumerate(records, 2):
        fill_hex = DOMAIN_COLORS.get(record.get("Category", ""), "FFFFFF")
        row_fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
 
        text   = record.get("Resume Text", "")
        values = [
            row_idx - 1,
            record.get("ID"),
            record.get("Category"),
            record.get("Quartile"),
            len(text),
            record.get("Completeness\n(out of 5)"),
            record.get("Gender Condition"),
            record.get("Injected Name"),
            text,
        ]
 
        for col_idx, value in enumerate(values, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = BORDER
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = (
                Alignment(wrap_text=True, vertical="top")
                if col_idx == 9
                else Alignment(horizontal="center", vertical="top")
            )
        ws.row_dimensions[row_idx].height = 80
 
 
def export_to_excel(records: list[dict], filepath: str) -> None:
    wb = openpyxl.Workbook()
 
    write_sheet(wb.active, records, "All CVs")
 
    for condition, title in [("male", "Male CVs"), ("female", "Female CVs")]:
        subset = [r for r in records if r.get("Gender Condition") == condition]
        write_sheet(wb.create_sheet(title), subset, title)
 
    wb.save(filepath)
    print(f"Exported to: {filepath}")
 

 
if __name__ == "__main__":
    print("=" * 60)
    print("HOBBIES INJECTION PIPELINE")
    print("=" * 60)
 
    print("\n[1/3] Loading Excel...")
    df      = pd.read_excel(INPUT_XLSX, sheet_name="All CVs")
    records = df.to_dict(orient="records")
    print(f"  Loaded {len(records)} rows")
 
    print("\n[2/3] Injecting hobbies (neutral CVs dropped)...")
    versioned = inject_hobbies(records)
 
    print("\n[3/3] Exporting to Excel...")
    export_to_excel(versioned, OUTPUT_XLSX)
 
    print("\nDone.")
