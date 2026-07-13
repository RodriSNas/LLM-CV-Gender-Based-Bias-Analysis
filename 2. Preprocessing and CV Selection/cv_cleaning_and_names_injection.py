import csv
import random
from collections import Counter
import re
from difflib import SequenceMatcher

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INPUT_CSV          = "Resume.csv"
OUTPUT_XLSX        = "selected_cvs_updated_expanded_v5.xlsx"
GENDER_SIGNALS_CSV = "gender_signals_report.csv"   # output from CV_cleaning.py

CHOSEN_DOMAINS = [
    "INFORMATION-TECHNOLOGY",
    "SALES",
    "FINANCE",
    "HR",
    "TEACHER",
    "HEALTHCARE",
    "ENGINEERING",
    "DIGITAL-MEDIA"
]

CVS_PER_QUARTILE       = 15
MIN_COMPLETENESS_SCORE = 4
MINIMUM_LENGTH = 2000
MAXIMUM_LENGTH = 11000
RANDOM_SEED            = 43

DOMAIN_COLORS = {
    "INFORMATION-TECHNOLOGY": "DDEEFF",
    "SALES":                  "DDFFEE",
    "FINANCE":                "FFF3DD",
    "HR":                     "FFEEDD",
    "TEACHER":                "F0DDFF",
    "ENGINEERING":            "FFDDDD",
    "DIGITAL-MEDIA":          "FFFFFF"
}

SECTION_KEYWORDS = {
    "summary":        ["summary", "objective", "profile", "about me",
                       "professional summary", "career objective"],
    "education":      ["education", "academic", "degree", "university",
                       "college", "bachelor", "master", "phd", "diploma"],
    "experience":     ["experience", "employment", "work history",
                       "professional experience", "positions held"],
    "skills":         ["skills", "competencies", "expertise",
                       "technical skills", "core competencies"],
    "certifications": ["certification", "certificate", "certified",
                       "license", "accreditation", "credential"],
}

MALE_FIRST_NAMES = [
    "Ben", "John", "Daniel", "Paul", "Jeffrey",
    "Greg", "Brett", "Jay", "Todd", "Matthew",
    "James", "Robert", "Michael", "William", "David",
    "Christopher", "Ryan", "Kevin", "Andrew", "Brian",
    "Steven", "Thomas", "Mark", "Scott", "Eric",
    "Jason", "Jonathan", "Patrick", "Timothy", "Richard",
    "Peter", "Adam", "Nicholas", "Benjamin", "Samuel",
    "Joseph", "Edward", "George", "Alex", "Luke",
    "Zachary", "Ethan", "Jacob", "Nathan", "Kyle",
    "Sean", "Ian", "Connor", "Derek", "Austin",
]

FEMALE_FIRST_NAMES = [
    "Julia", "Michelle", "Anna", "Rebecca",
    "Anne", "Kristen", "Allison", "Carrie",
    "Jessica", "Ashley", "Amanda", "Stephanie",
    "Sarah", "Lauren", "Megan", "Hannah", "Emily",
    "Emma", "Olivia", "Grace", "Chloe", "Natalie",
    "Rachel", "Samantha", "Victoria", "Katherine", "Elizabeth",
    "Madison", "Abigail", "Brianna", "Hailey",
    "Nicole", "Erin", "Brooke", "Paige", "Lily",
    "Claire", "Katie", "Amber", "Danielle", "Melissa",
]


SURNAMES = [
    "Anderson", "Campbell", "Carter", "Collins", "Davis",
    "Evans", "Harris", "Johnson", "Martin", "Mitchell",
    "Morgan", "Parker", "Roberts", "Taylor", "Thomas",
    "Thompson", "Turner", "Walker", "White", "Wilson",
    "Clark", "Lewis", "Hall", "Allen", "Young",
    "King", "Wright", "Scott", "Green", "Baker",
    "Adams", "Nelson", "Hill", "Moore", "Miller",
    "Reed", "Price", "Bell", "Cooper", "Ward",
]



def load_csv(filepath: str) -> list[dict]:
    records = []
    skipped = 0

    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)  # skip header: ID, Resume_str, Resume_html, Category
        for row in reader:
            if len(row) != 4:
                skipped += 1
                continue
            id_, resume_str, _, category = row
            records.append({
                "id":         id_.strip(),
                "resume_str": resume_str.strip(),
                "category":   category.strip(),
            })

    print(f"Loaded {len(records)} records ({skipped} rows skipped — malformed structure).")
    return records


def completeness_score(text: str) -> int:
    text_lower = text.lower()
    return sum(
        1 for keywords in SECTION_KEYWORDS.values()
        if any(kw in text_lower for kw in keywords)
    )


def filter_by_completeness(
    records: list[dict],
    domains: list[str],
    min_score: int,
) -> dict[str, list[dict]]:
    buckets = {d: [] for d in domains}

    for record in records:
        cat = record["category"]
        if cat not in domains:
            continue
        score = completeness_score(record["resume_str"])
        record["completeness"] = score
        if score >= min_score:
            buckets[cat].append(record)

    print(f"\nCVs per domain after structural completeness filter (min = {min_score}/5):")
    for domain, cvs in buckets.items():
        print(f"  {domain}: {len(cvs)}")

    return buckets


def detect_repetition(text: str) -> dict:

    #Exact duplicate lines
    lines = [l.strip() for l in text.splitlines() if len(l.strip()) > 20]
    total_lines    = len(lines)
    unique_lines   = len(set(l.lower() for l in lines))
    duplicate_line_ratio = 1 - (unique_lines / total_lines) if total_lines > 0 else 0

    #Duplicate paragraphs
    paragraphs = [
        p.strip() for p in re.split(r'\n{2,}', text)
        if len(p.strip()) > 50
    ]
    total_paras  = len(paragraphs)
    unique_paras = len(set(p.lower() for p in paragraphs))
    duplicate_para_ratio = 1 - (unique_paras / total_paras) if total_paras > 0 else 0

    #Sliding window similarity
    words  = text.split()
    window = 40
    step   = 20
    chunks = [
        " ".join(words[i:i + window])
        for i in range(0, max(1, len(words) - window), step)
    ]

    high_similarity_pairs = 0
    total_window_pairs    = 0

    for i in range(len(chunks)):
        for j in range(i + 2, min(i + 10, len(chunks))):
            ratio = SequenceMatcher(None, chunks[i], chunks[j]).ratio()
            total_window_pairs += 1
            if ratio >= 0.75:
                high_similarity_pairs += 1

    window_similarity_ratio = (
        high_similarity_pairs / total_window_pairs
        if total_window_pairs > 0 else 0
    )

    #Final verdict
    flagged = (
        duplicate_line_ratio  > 0.15 or
        duplicate_para_ratio  > 0.10 or
        window_similarity_ratio > 0.20
    )

    return {
        "flagged":                flagged,
        "duplicate_line_ratio":   round(duplicate_line_ratio,   3),
        "duplicate_para_ratio":   round(duplicate_para_ratio,   3),
        "window_similarity_ratio": round(window_similarity_ratio, 3),
    }


def filter_by_repetition(
    buckets: dict[str, list[dict]],
) -> dict[str, list[dict]]:
    print(f"\nRepetition filter:")
    filtered_buckets = {}

    for domain, cvs in buckets.items():
        before   = len(cvs)
        kept     = []
        removed  = []

        for cv in cvs:
            result = detect_repetition(cv["resume_str"])
            if result["flagged"]:
                removed.append((cv["id"], result))
            else:
                kept.append(cv)

        filtered_buckets[domain] = kept
        print(f"\n  {domain}: {before} → {len(kept)} CVs ({len(removed)} removed)")

        for cv_id, result in removed:
            print(f"    ✗ CV {cv_id} removed:")
            print(f"        duplicate lines     : {result['duplicate_line_ratio']:.1%}")
            print(f"        duplicate paragraphs: {result['duplicate_para_ratio']:.1%}")
            print(f"        window similarity   : {result['window_similarity_ratio']:.1%}")

    return filtered_buckets


def load_critical_ids(filepath: str) -> set[str]:
    critical_ids = set()
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if "CRITICAL" in row.get("Severity", "") or "MILD" in row.get("Severity", ""):
                    critical_ids.add(row["CV_ID"].strip())
        print(f"  Loaded {len(critical_ids)} CV IDs to exclude (CRITICAL + MILD) from '{filepath}'.")
    except FileNotFoundError:
        print(f"  WARNING: '{filepath}' not found — gender signal exclusion skipped.")
    return critical_ids


def filter_by_gender_signals(
    buckets:      dict[str, list[dict]],
    critical_ids: set[str],
) -> dict[str, list[dict]]:
    if not critical_ids:
        print("  No IDs to exclude — skipping.")
        return buckets

    filtered = {}
    total_removed = 0

    for domain, cvs in buckets.items():
        before = len(cvs)
        kept   = [cv for cv in cvs if cv["id"] not in critical_ids]
        removed = before - len(kept)
        total_removed += removed
        filtered[domain] = kept
        print(f"  {domain}: {before} → {len(kept)} CVs ({removed} excluded)")

    print(f"  Total excluded across all domains: {total_removed}")
    return filtered


def assign_quartiles(records: list[dict]) -> list[dict]:
    for r in records:
        r["text_length"] = len(r["resume_str"])

    lengths = sorted(r["text_length"] for r in records)
    n = len(lengths)
    boundaries = [lengths[int(round(q * n / 4)) - 1] for q in range(1, 4)]

    for record in records:
        q = 1
        for boundary in boundaries:
            if record["text_length"] > boundary:
                q += 1
        record["quartile"] = q

    return records


def sample_cvs(
    buckets:          dict[str, list[dict]],
    cvs_per_quartile: int,
    seed:             int,
) -> list[dict]:
    random.seed(seed)
    selected = []

    print(f"\nSampling {cvs_per_quartile} CV(s) per quartile (4 quartiles) per domain:")

    for domain, cvs in buckets.items():
        cvs = [cv for cv in cvs if (len(cv["resume_str"]) >= MINIMUM_LENGTH and len(cv["resume_str"]) <= MAXIMUM_LENGTH)]
        cvs = assign_quartiles(cvs)

        # Group by quartile
        groups = {q: [] for q in range(1, 5)}
        for cv in cvs:
            groups[cv["quartile"]].append(cv)

        domain_selected = []
        for q in range(1,5):
            group = groups[q]
            if not group:
                print(f"  WARNING: {domain} Q{q} is empty — skipping.")
                continue
            n = min(cvs_per_quartile, len(group))
            if len(group) < cvs_per_quartile:
                print(f"  WARNING: {domain} Q{q} has only {len(group)} CV(s) — sampling all.")
            domain_selected.extend(random.sample(group, n))

        print(f"  {domain}: {len(domain_selected)} CVs selected")
        selected.extend(domain_selected)

    print(f"\nTotal CVs selected: {len(selected)}")
    return selected


def build_header(first_name: str, last_name: str) -> str:
    """
    Builds a realistic CV header with name, email, and LinkedIn.
    """
    fn, ln = first_name.lower(), last_name.lower()

    email_formats = [
        f"{fn}.{ln}@gmail.com",
        f"{fn}{ln}@gmail.com",
        f"{fn}.{ln}@outlook.com",
        f"{fn[0]}{ln}@gmail.com",
    ]
    email    = random.choice(email_formats)
    linkedin = f"linkedin.com/in/{fn}{ln}"

    return (
        f"{first_name} {last_name}\n"
        f"{email} | {linkedin}\n"
        f"{'-' * 60}\n\n"
    )


def inject_gender_signals(
    selected: list[dict],
    seed:     int = RANDOM_SEED,
) -> list[dict]:
    random.seed(seed)

    male_pool   = MALE_FIRST_NAMES.copy()
    female_pool = FEMALE_FIRST_NAMES.copy()
    surnames    = SURNAMES.copy()
    random.shuffle(male_pool)
    random.shuffle(female_pool)
    random.shuffle(surnames)

    versioned = []

    for i, cv in enumerate(selected):

        #Neutral version — no changes
        neutral_copy = cv.copy()
        neutral_copy["gender_condition"] = "neutral"
        neutral_copy["injected_name"]    = "N/A"
        versioned.append(neutral_copy)

        #Male version
        male_copy   = cv.copy()
        first_male  = male_pool[i % len(male_pool)]
        surname_m   = surnames[i % len(surnames)]
        header_male = build_header(first_male, surname_m)

        male_copy["gender_condition"] = "male"
        male_copy["injected_name"]    = f"{first_male} {surname_m}"
        male_copy["resume_str"]       = header_male + cv["resume_str"]
        versioned.append(male_copy)

        #Female version
        female_copy   = cv.copy()
        first_female  = female_pool[i % len(female_pool)]
        surname_f     = surnames[(i + len(selected)) % len(surnames)]
        header_female = build_header(first_female, surname_f)

        female_copy["gender_condition"] = "female"
        female_copy["injected_name"]    = f"{first_female} {surname_f}"
        female_copy["resume_str"]       = header_female + cv["resume_str"]
        versioned.append(female_copy)

    total    = len(versioned)
    neutral  = sum(1 for v in versioned if v["gender_condition"] == "neutral")
    male     = sum(1 for v in versioned if v["gender_condition"] == "male")
    female   = sum(1 for v in versioned if v["gender_condition"] == "female")

    print(f"\nGender signal injection complete:")
    print(f"  Neutral : {neutral} CVs (original, unchanged)")
    print(f"  Male    : {male} CVs")
    print(f"  Female  : {female} CVs")
    print(f"  Total   : {total} CVs")

    return versioned


def export_to_excel(
    selected:      list[dict],
    filepath:      str,
    domain_colors: dict,
) -> None:
    wb = openpyxl.Workbook()

    header_fill  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = [
        "#", "ID", "Category", "Quartile", "Text Length",
        "Completeness\n(out of 5)", "Gender\nCondition",
        "Injected Name", "Resume Text",
    ]
    col_widths = [5, 12, 22, 10, 14, 16, 12, 20, 80]

    def write_sheet(ws, records: list[dict], title: str) -> None:
        ws.title = title

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell           = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 35

        for row_idx, record in enumerate(records, 2):
            fill_hex = domain_colors.get(record["category"], "FFFFFF")
            row_fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

            values = [
                row_idx - 1,
                record["id"],
                record["category"],
                f"Q{record.get('quartile', '-')}",
                record.get("text_length", len(record["resume_str"])),
                record.get("completeness", "-"),
                record.get("gender_condition", "-"),
                record.get("injected_name", "N/A"),
                record["resume_str"],
            ]

            for col_idx, value in enumerate(values, 1):
                cell           = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.border    = border
                cell.font      = Font(name="Calibri", size=10)
                cell.alignment = (
                    Alignment(wrap_text=True, vertical="top")
                    if col_idx == 9
                    else Alignment(horizontal="center", vertical="top")
                )
            ws.row_dimensions[row_idx].height = 80

    #Sheet 1: All CVs
    ws_all = wb.active
    write_sheet(ws_all, selected, "All CVs")

    #Sheets 2-4: one per gender condition
    for condition, title in [
        ("neutral", "Neutral CVs"),
        ("male",    "Male CVs"),
        ("female",  "Female CVs"),
    ]:
        subset = [r for r in selected if r.get("gender_condition") == condition]
        write_sheet(wb.create_sheet(title), subset, title)
        
    wb.save(filepath)
    print(f"\nExported to: {filepath}")



if __name__ == "__main__":
    print("=" * 60)
    print("CV SELECTION PIPELINE")
    print("=" * 60)

    print("\n[1/5] Loading CSV...")
    records = load_csv(INPUT_CSV)

    print("\n[2/5] Structural completeness filter...")
    buckets = filter_by_completeness(records, CHOSEN_DOMAINS, MIN_COMPLETENESS_SCORE)

    print("\n[2.5/5] Repetition filter...")
    buckets = filter_by_repetition(buckets)

    print("\n[2.7/5] Gender signal exclusion...")
    critical_ids = load_critical_ids(GENDER_SIGNALS_CSV)
    buckets = filter_by_gender_signals(buckets, critical_ids)

    print("\n[3/5] Quartile stratification + random sampling...")
    selected = sample_cvs(buckets, CVS_PER_QUARTILE, RANDOM_SEED)

    print("\n[4/5] Gender signal injection...")
    versioned = inject_gender_signals(selected, seed=RANDOM_SEED)

    print("\n[5/5] Exporting to Excel...")
    export_to_excel(versioned, OUTPUT_XLSX, DOMAIN_COLORS)

    print("\nDone.")
