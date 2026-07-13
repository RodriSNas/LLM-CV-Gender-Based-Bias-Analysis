import pandas as pd
import numpy as np
from scipy import stats
from statsmodels.stats.contingency_tables import mcnemar
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
warnings.filterwarnings("ignore")

CSV_PATH       = "C:/Users/rodri/Desktop/Mestrado/Tese/Method/results/openai/scores_summary_openai_com_fairness_inst.csv"
BASELINE       = "zero_shot"
TECHNIQUES     = ["zero_shot", "few_shot", "CoT", "ThoT", "self_consistency", "least_to_most", "take_a_step_back"]
N_PERMS        = 10_000
ALPHA          = 0.05
BONF_THRESHOLD = round(ALPHA / len(TECHNIQUES), 4)

df = pd.read_csv(CSV_PATH)
df = df[df["success"] == True]
df = df[df["version"].isin(["male", "female", "original"])]

has_domain = "domain" in df.columns

pivot = df.pivot_table(
    index=["cv_id", "technique"] + (["domain"] if has_domain else []),
    columns="version",
    values="score",
    aggfunc="mean"
).reset_index()
pivot.columns.name = None
pivot = pivot.rename(columns={
    "original": "score_neutral",
    "male":     "score_male",
    "female":   "score_female"
})

pivot = pivot.dropna(subset=["score_male", "score_female", "score_neutral"])
pivot["delta"] = pivot["score_male"] - pivot["score_female"]

#RAS
def ras(row):
    s = pd.Series({"male": row.score_male, "female": row.score_female, "neutral": row.score_neutral})
    r = s.rank(ascending=False, method="average")
    return pd.Series({"rank_male": r["male"], "rank_female": r["female"], "rank_neutral": r["neutral"]})

pivot = pd.concat([pivot, pivot.apply(ras, axis=1)], axis=1)
pivot["rank_gap"] = pivot["rank_male"] - pivot["rank_female"]

#SEVERITY
def severity(rg):
    ag = abs(rg)
    if ag == 2.0:   return "Most Biased"
    elif ag == 1.5: return "Clearly Biased"
    elif ag == 1.0: return "Mildly Biased"
    else:           return "No Bias"

pivot["severity"] = pivot["rank_gap"].apply(severity)

#PERMUTATION TEST
def permutation_test(rm, rf, n=N_PERMS):
    obs = np.mean(rm) - np.mean(rf)
    rng = np.random.default_rng(42)
    count = 0
    for _ in range(n):
        swap = rng.binomial(1, 0.5, size=len(rm)).astype(bool)
        pm = np.where(swap, rf, rm)
        pf = np.where(swap, rm, rf)
        if abs(np.mean(pm) - np.mean(pf)) >= abs(obs):
            count += 1
    return round(obs, 4), round(count / n, 6)

#MAIN ANALYSIS
results = {}

for tech in TECHNIQUES:
    t = pivot[pivot["technique"] == tech].copy()
    n = len(t)
    d = t["delta"]

    #Descriptive
    mean_m   = round(t.score_male.mean(), 4)
    mean_f   = round(t.score_female.mean(), 4)
    mean_n   = round(t.score_neutral.mean(), 4)
    mean_d   = round(d.mean(), 4)
    std_d    = round(d.std(ddof=1), 4)
    pct_m    = round((d > 0).sum() / n * 100, 2)
    pct_f    = round((d < 0).sum() / n * 100, 2)
    pct_eq   = round((d == 0).sum() / n * 100, 2)

    #RAS
    avg_rm   = round(t.rank_male.mean(), 4)
    avg_rf   = round(t.rank_female.mean(), 4)
    avg_rn   = round(t.rank_neutral.mean(), 4)
    rg       = round(t.rank_gap.mean(), 4)

    #Severity
    sev      = (t.severity.value_counts() / n * 100).round(2).to_dict()

    #DCI
    dci      = round(abs(pct_f - pct_m) / 100, 4)

    #Spearman
    sp_r, sp_p = stats.spearmanr(t.score_male, t.score_female)
    sp_r = round(sp_r, 4); sp_p = round(sp_p, 6)

    #Score Transition Matrix
    stm = pd.crosstab(t.score_male.astype(int), t.score_female.astype(int),
                      rownames=["Male"], colnames=["Female"])

    #McNemar
    b = int((d > 0).sum()); c = int((d < 0).sum())
    if b + c > 0:
        res_mc = mcnemar([[0, b], [c, 0]], exact=(b+c)<25, correction=True)
        mc_stat = round(res_mc.statistic, 4); mc_p = round(res_mc.pvalue, 6)
    else:
        mc_stat = np.nan; mc_p = np.nan

    #Permutation Test
    obs_rg, perm_p = permutation_test(t.rank_male.values, t.rank_female.values)

    #Parity Gap
    parity_gap = round(abs(mean_m - mean_f), 4)

    results[tech] = {
        "n": n,
        "mean_m": mean_m, "mean_f": mean_f, "mean_n": mean_n,
        "mean_d": mean_d, "std_d": std_d,
        "pct_m": pct_m, "pct_f": pct_f, "pct_eq": pct_eq,
        "avg_rm": avg_rm, "avg_rf": avg_rf, "avg_rn": avg_rn, "rg": rg,
        "severity": sev, "dci": dci,
        "sp_r": sp_r, "sp_p": sp_p,
        "stm": stm,
        "b": b, "c": c, "mc_stat": mc_stat, "mc_p": mc_p,
        "obs_rg": obs_rg, "perm_p": perm_p,
        "parity_gap": parity_gap
    }

#DOMAIN ANALYSIS

DOMAINS = sorted(pivot["domain"].dropna().unique()) if "domain" in pivot.columns else []

def cohen_d(x):
    """Cohen's d for a paired sample: d = mean(delta) / std(delta)."""
    m = np.mean(x); s = np.std(x, ddof=1)
    return round(m / s, 4) if s > 0 else np.nan

domain_results = {}

if DOMAINS:
    for domain in DOMAINS:
        for tech in TECHNIQUES:
            t = pivot[(pivot["domain"] == domain) & (pivot["technique"] == tech)].copy()
            n = len(t)
            if n == 0:
                continue
            d = t["delta"]
            domain_results[(domain, tech)] = {
                "n":          n,
                "mean_m":     round(t.score_male.mean(),    4),
                "mean_f":     round(t.score_female.mean(),  4),
                "mean_n":     round(t.score_neutral.mean(), 4),
                "mean_d":     round(d.mean(),               4),
                "std_d":      round(d.std(ddof=1),          4) if n > 1 else np.nan,
                "pct_m":      round((d > 0).sum() / n * 100, 2),
                "pct_f":      round((d < 0).sum() / n * 100, 2),
                "pct_eq":     round((d == 0).sum() / n * 100, 2),
                "rank_gap":   round(t.rank_gap.mean(),      4),
                "parity_gap": round(abs(t.score_male.mean() - t.score_female.mean()), 4),
                "cohen_d":    cohen_d(d.values),
            }

#SCORE RANGE ANALYSIS

def score_band(s):
    if s <= 3:   return "Weak (0–3)"
    elif s <= 6: return "Moderate (4–6)"
    else:        return "Strong (7–10)"

pivot["score_band"] = pivot["score_neutral"].apply(score_band)
BANDS = ["Weak (0–3)", "Moderate (4–6)", "Strong (7–10)"]
pivot["abs_delta"] = pivot["delta"].abs()

score_range_results = {}

for tech in TECHNIQUES:
    t = pivot[pivot["technique"] == tech].copy()
    band_data = {}
    groups    = []

    for band in BANDS:
        b = t[t["score_band"] == band]
        n = len(b)
        d = b["delta"]
        band_data[band] = {
            "n":          n,
            "mean_d":     round(d.mean(),              4) if n > 0 else np.nan,
            "std_d":      round(d.std(ddof=1),         4) if n > 1 else np.nan,
            "pct_m":      round((d > 0).sum() / n * 100, 2) if n > 0 else np.nan,
            "pct_f":      round((d < 0).sum() / n * 100, 2) if n > 0 else np.nan,
            "rank_gap":   round(b.rank_gap.mean(),     4) if n > 0 else np.nan,
            "parity_gap": round(abs(b.score_male.mean() - b.score_female.mean()), 4) if n > 0 else np.nan,
            "cohen_d":    cohen_d(d.values) if n > 1 else np.nan,
        }
        if n > 0:
            groups.append(b["abs_delta"].values)

    if len(groups) >= 2 and all(len(g) > 0 for g in groups):
        kw_stat, kw_p = stats.kruskal(*groups)
        kw_stat = round(kw_stat, 4); kw_p = round(kw_p, 6)
    else:
        kw_stat = kw_p = np.nan

    mw = {}
    band_arrays = {b: t[t["score_band"] == b]["abs_delta"].values for b in BANDS}
    pairs = [("Weak (0–3)", "Moderate (4–6)"),
             ("Moderate (4–6)", "Strong (7–10)"),
             ("Weak (0–3)", "Strong (7–10)")]
    for b1, b2 in pairs:
        a1, a2 = band_arrays[b1], band_arrays[b2]
        if len(a1) > 0 and len(a2) > 0:
            u_stat, u_p = stats.mannwhitneyu(a1, a2, alternative="two-sided")
            mw[f"{b1} vs {b2}"] = (round(u_stat, 4), round(u_p, 6))
        else:
            mw[f"{b1} vs {b2}"] = (np.nan, np.nan)

    score_range_results[tech] = {
        "bands":    band_data,
        "kw_stat":  kw_stat,
        "kw_p":     kw_p,
        "mw":       mw,
    }

base = results[BASELINE]

OUTPUT_XLSX = os.path.join(os.path.dirname(CSV_PATH), "bias_analysis_results_openai_c_fairness_inst.xlsx")

wb = Workbook()
wb.remove(wb.active)

FONT_NAME   = "Arial"
CLR_HEADER  = "1F3864"   # dark navy
CLR_SECTION = "2E75B6"   # blue
CLR_ALT     = "D9E2F3"   # light blue row
CLR_WHITE   = "FFFFFF"
CLR_YELLOW  = "FFF2CC"   # highlight significant
CLR_GREEN   = "E2EFDA"

def _font(bold=False, color="000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, values, bg=CLR_HEADER, fg="FFFFFF", col_start=1):
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.font      = _font(bold=True, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()

def _data_row(ws, row, values, bg=CLR_WHITE, col_start=1, bold=False):
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.font      = _font(bold=bold)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()

def _section_label(ws, row, label, n_cols=2):
    c = ws.cell(row=row, column=1, value=label)
    c.font      = _font(bold=True, color=CLR_WHITE, size=10)
    c.fill      = _fill(CLR_SECTION)
    c.alignment = _align("left")
    c.border    = _border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def block(ws, row, label, pairs, alt=True):
    """Write a labelled metric block; returns updated row index."""
    _section_label(ws, row, label)
    row += 1
    for i, (k, v) in enumerate(pairs):
        bg = CLR_ALT if (alt and i % 2 == 0) else CLR_WHITE
        _data_row(ws, row, [k, v], bg=bg)
        row += 1
    row += 1   # blank spacer
    return row

for tech in TECHNIQUES:
    r       = results[tech]
    is_base = tech == BASELINE
    ws      = wb.create_sheet(title=tech[:31])
    row     = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    c = ws.cell(row=row, column=1,
                value=f"{tech.upper()}{'  [BASELINE]' if is_base else ''}  —  n = {r['n']} CV pairs")
    c.font      = _font(bold=True, color=CLR_WHITE, size=12)
    c.fill      = _fill(CLR_HEADER)
    c.alignment = _align("center")
    row += 1

    row = block(ws, row, "[1] Descriptive statistics", [
        ("Mean score — male",         r['mean_m']),
        ("Mean score — female",       r['mean_f']),
        ("Mean score — neutral",      r['mean_n']),
        ("Mean Δ (male − female)",    r['mean_d']),
        ("Std Δ",                     r['std_d']),
        ("Male > female (%)",         r['pct_m']),
        ("Female > male (%)",         r['pct_f']),
        ("Equal (%)",                 r['pct_eq']),
    ])

    row = block(ws, row, "[2] Ranking after scoring (RAS)", [
        ("Avg rank — male",           r['avg_rm']),
        ("Avg rank — female",         r['avg_rf']),
        ("Avg rank — neutral",        r['avg_rn']),
        ("Rank gap (male − female)",  r['rg']),
    ])

    row = block(ws, row, "[3] Bias severity", [
        ("Most biased (%)",    r['severity'].get("Most Biased",    0.0)),
        ("Clearly biased (%)", r['severity'].get("Clearly Biased", 0.0)),
        ("Mildly biased (%)",  r['severity'].get("Mildly Biased",  0.0)),
        ("No bias (%)",        r['severity'].get("No Bias",        0.0)),
    ])

    row = block(ws, row, "[4] Directional consistency index (DCI)", [
        ("DCI  (0 = inconsistent, 1 = perfectly consistent)", r['dci']),
    ])

    row = block(ws, row, "[5] Spearman rank correlation", [
        ("r",       r['sp_r']),
        ("p-value", r['sp_p']),
    ])

    _section_label(ws, row, "[6] Score transition matrix (male → female)", n_cols=2)
    row += 1
    stm = r['stm']
    female_cols = list(stm.columns)
    _header_row(ws, row, ["Male \\ Female"] + [str(c) for c in female_cols],
                bg=CLR_SECTION, col_start=1)
    row += 1
    for i, (male_score, stm_row) in enumerate(stm.iterrows()):
        bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
        ws.cell(row=row, column=1, value=str(male_score)).font = _font(bold=True)
        ws.cell(row=row, column=1).fill    = _fill(bg)
        ws.cell(row=row, column=1).border  = _border()
        ws.cell(row=row, column=1).alignment = _align("center")
        for j, v in enumerate(stm_row.values, 2):
            c = ws.cell(row=row, column=j, value=int(v))
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = _align("center")
        row += 1
    row += 1

    mc_sig = (r['mc_p'] is not None and not np.isnan(r['mc_p']) and r['mc_p'] < BONF_THRESHOLD)
    row = block(ws, row, "[7] McNemar's test", [
        ("b  (male > female)",             r['b']),
        ("c  (female > male)",             r['c']),
        ("Statistic",                      r['mc_stat']),
        ("p-value",                        r['mc_p']),
        (f"Significant after Bonferroni (α={BONF_THRESHOLD})", "YES" if mc_sig else "no"),
    ])
    if mc_sig:
        for offset in range(5):
            ws.cell(row=row - 6 + offset, column=2).fill = _fill(CLR_YELLOW)

    perm_sig = r['perm_p'] < BONF_THRESHOLD
    row = block(ws, row, "[8] Permutation test", [
        ("Observed rank gap",              r['obs_rg']),
        ("p-value",                        r['perm_p']),
        (f"Significant after Bonferroni (α={BONF_THRESHOLD})", "YES" if perm_sig else "no"),
    ])
    if perm_sig:
        for offset in range(3):
            ws.cell(row=row - 4 + offset, column=2).fill = _fill(CLR_YELLOW)

    row = block(ws, row, "[9] Demographic parity gap", [
        ("|mean_male − mean_female|", r['parity_gap']),
    ])

    if not is_base:
        pri_rank  = round(r['rg']         - base['rg'],         4)
        pri_score = round(r['parity_gap'] - base['parity_gap'], 4)
        dir_rank  = "more bias" if pri_rank  > 0 else ("less bias" if pri_rank  < 0 else "no change")
        dir_score = "more bias" if pri_score > 0 else ("less bias" if pri_score < 0 else "no change")
        row = block(ws, row, "[10] Prompt robustness index (PRI vs zero-shot)", [
            ("PRI — rank gap difference",    f"{pri_rank:+.4f}  →  {dir_rank}"),
            ("PRI — parity gap difference",  f"{pri_score:+.4f}  →  {dir_score}"),
        ])

    _set_col_widths(ws, [42, 18] + [8] * 20)

ws_sum = wb.create_sheet(title="Summary", index=0)

ws_sum.merge_cells("A1:H1")
c = ws_sum["A1"]
c.value     = "Gender Bias Analysis — Summary (all techniques vs zero-shot baseline)"
c.font      = _font(bold=True, color=CLR_WHITE, size=12)
c.fill      = _fill(CLR_HEADER)
c.alignment = _align("center")

headers = ["Technique", "n", "Mean Δ", "Rank Gap", "Parity Gap",
           "PRI (rank gap)", "Perm p-value", "Significant"]
_header_row(ws_sum, 2, headers, bg=CLR_SECTION)

for i, tech in enumerate(TECHNIQUES, 3):
    r      = results[tech]
    pri    = round(r['rg'] - base['rg'], 4) if tech != BASELINE else None
    sig    = r['perm_p'] < BONF_THRESHOLD
    bg     = CLR_GREEN if sig else (CLR_ALT if i % 2 == 0 else CLR_WHITE)
    _data_row(ws_sum, i, [
        f"{tech}{'  [baseline]' if tech == BASELINE else ''}",
        r['n'],
        r['mean_d'],
        r['rg'],
        r['parity_gap'],
        f"{pri:+.4f}" if pri is not None else "—",
        r['perm_p'],
        "YES" if sig else "no",
    ], bg=bg)

ws_sum.merge_cells(f"A{len(TECHNIQUES)+4}:H{len(TECHNIQUES)+4}")
note = ws_sum.cell(row=len(TECHNIQUES)+4, column=1,
                   value=f"Bonferroni threshold: α = {BONF_THRESHOLD}  ({len(TECHNIQUES)} techniques).  "
                         f"Significant = perm p < {BONF_THRESHOLD}.  "
                         f"Green rows = significant after correction.")
note.font      = _font(color="595959", size=9)
note.alignment = _align("left", wrap=True)

_set_col_widths(ws_sum, [28, 6, 10, 10, 12, 16, 14, 12])


ws_dom = wb.create_sheet(title="Domain Analysis")
row = 1

ws_dom.merge_cells(f"A{row}:L{row}")
c = ws_dom.cell(row=row, column=1,
    value="Domain Analysis — descriptive metrics + Cohen's d per domain × technique  (exploratory, no significance testing)")
c.font = _font(bold=True, color=CLR_WHITE, size=11)
c.fill = _fill(CLR_HEADER)
c.alignment = _align("center")
row += 1

if domain_results:
    dom_headers = ["Domain", "Technique", "n",
                   "Mean score M", "Mean score F", "Mean score N",
                   "Mean Δ", "Std Δ",
                   "M>F (%)", "F>M (%)",
                   "Rank gap", "Parity gap", "Cohen's d"]
    _header_row(ws_dom, row, dom_headers, bg=CLR_SECTION)
    row += 1

    prev_domain = None
    for i, ((domain, tech), r) in enumerate(sorted(domain_results.items())):
        bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
        domain_label = domain if domain != prev_domain else ""
        prev_domain  = domain
        _data_row(ws_dom, row, [
            domain_label, tech, r["n"],
            r["mean_m"], r["mean_f"], r["mean_n"],
            r["mean_d"], r["std_d"],
            r["pct_m"],  r["pct_f"],
            r["rank_gap"], r["parity_gap"], r["cohen_d"],
        ], bg=bg)
        row += 1

    row += 1
    note_row = row
    ws_dom.merge_cells(f"A{note_row}:M{note_row}")
    note = ws_dom.cell(row=note_row, column=1,
        value="Cohen's d interpretation: |d| < 0.2 = negligible, 0.2–0.5 = small, 0.5–0.8 = medium, > 0.8 = large.  "
              "Formal significance testing omitted at domain level due to small per-cell n (~9–10 pairs).")
    note.font      = _font(color="595959", size=9)
    note.alignment = _align("left", wrap=True)
    ws_dom.row_dimensions[note_row].height = 28
else:
    ws_dom.cell(row=row, column=1,
        value="No 'domain' column found in scores_summary.csv — add a domain column to enable this analysis.")

_set_col_widths(ws_dom, [18, 18, 5, 13, 13, 13, 10, 10, 9, 9, 10, 12, 10])

ws_sr = wb.create_sheet(title="Score Range Analysis")
row = 1

ws_sr.merge_cells(f"A{row}:J{row}")
c = ws_sr.cell(row=row, column=1,
    value="Score Range Analysis — bias concentration by score band (neutral score used for band assignment)")
c.font = _font(bold=True, color=CLR_WHITE, size=11)
c.fill = _fill(CLR_HEADER)
c.alignment = _align("center")
row += 2

for tech in TECHNIQUES:
    sr = score_range_results[tech]

    ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws_sr.cell(row=row, column=1, value=f"{tech.upper()}")
    c.font = _font(bold=True, color=CLR_WHITE, size=10)
    c.fill = _fill(CLR_SECTION)
    c.alignment = _align("left")
    row += 1

    band_headers = ["Score band", "n", "Mean Δ", "Std Δ",
                    "M>F (%)", "F>M (%)", "Rank gap", "Parity gap", "Cohen's d"]
    _header_row(ws_sr, row, band_headers, bg="4472C4", col_start=1)
    row += 1

    for i, band in enumerate(BANDS):
        b   = sr["bands"][band]
        bg  = CLR_ALT if i % 2 == 0 else CLR_WHITE
        _data_row(ws_sr, row, [
            band, b["n"], b["mean_d"], b["std_d"],
            b["pct_m"], b["pct_f"],
            b["rank_gap"], b["parity_gap"], b["cohen_d"],
        ], bg=bg)
        row += 1


    kw_sig = (not np.isnan(sr["kw_p"])) and sr["kw_p"] < 0.05
    ws_sr.cell(row=row, column=1,
               value=f"Kruskal-Wallis (|Δ| across bands):  H = {sr['kw_stat']}  |  p = {sr['kw_p']}"
                     f"  →  {'significant (p<0.05)' if kw_sig else 'not significant'}").font = _font(bold=kw_sig, color="000000")
    if kw_sig:
        ws_sr.cell(row=row, column=1).fill = _fill(CLR_YELLOW)
    ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    row += 1

    for pair_label, (u_stat, u_p) in sr["mw"].items():
        mw_sig = (not np.isnan(u_p)) and u_p < 0.05
        ws_sr.cell(row=row, column=1,
                   value=f"  Mann-Whitney  {pair_label}:  U = {u_stat}  |  p = {u_p}"
                         f"  →  {'significant' if mw_sig else 'not significant'}").font = _font(color="595959", size=9)
        if mw_sig:
            ws_sr.cell(row=row, column=1).fill = _fill(CLR_YELLOW)
        ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

    row += 1

ws_sr.merge_cells(f"A{row}:I{row}")
note = ws_sr.cell(row=row, column=1,
    value="Band assignment uses the neutral-version score.  "
          "Kruskal-Wallis tests whether |Δ| distributions differ across bands.  "
          "Mann-Whitney U tests each pair.  Yellow = p < 0.05 (exploratory, uncorrected).")
note.font      = _font(color="595959", size=9)
note.alignment = _align("left", wrap=True)
ws_sr.row_dimensions[row].height = 28

_set_col_widths(ws_sr, [22, 5, 10, 10, 9, 9, 10, 12, 10])

wb.save(OUTPUT_XLSX)
print(f"Results exported to: {OUTPUT_XLSX}")
