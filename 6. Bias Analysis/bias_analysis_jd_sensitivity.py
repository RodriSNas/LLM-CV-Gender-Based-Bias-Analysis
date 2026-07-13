import os
import pandas as pd
import numpy as np
from scipy import stats
from statsmodels.stats.contingency_tables import mcnemar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

CSV_PATH   = "C:/Users/rodri/Desktop/Mestrado/Tese/Method/results/summary_domain_analysis.csv"
OUTPUT_XLSX = os.path.join(os.path.dirname(CSV_PATH), "bias_analysis_domain_analysis.xlsx")

BASELINE   = "zero_shot"
TECHNIQUES = ["zero_shot", "few_shot", "CoT", "ThoT", "self_consistency", "least_to_most", "take_a_step_back"]
N_PERMS    = 10_000
ALPHA      = 0.05
BONF_THRESHOLD = round(ALPHA / len(TECHNIQUES), 4)
BANDS      = ["Weak (0–3)", "Moderate (4–6)", "Strong (7–10)"]

FONT_NAME   = "Arial"
CLR_HEADER  = "1F3864"
CLR_SECTION = "2E75B6"
CLR_ALT     = "D9E2F3"
CLR_WHITE   = "FFFFFF"
CLR_YELLOW  = "FFF2CC"
CLR_GREEN   = "E2EFDA"

def _font(bold=False, color="000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_row(ws, row, values, bg=CLR_HEADER, fg="FFFFFF", col_start=1):
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _font(bold=True, color=fg); c.fill = _fill(bg)
        c.alignment = _align("center"); c.border = _border()

def _data_row(ws, row, values, bg=CLR_WHITE, col_start=1, bold=False):
    for i, v in enumerate(values, col_start):
        c = ws.cell(row=row, column=i, value=v)
        c.font = _font(bold=bold); c.fill = _fill(bg)
        c.alignment = _align("center"); c.border = _border()

def _section_label(ws, row, label, n_cols=2):
    c = ws.cell(row=row, column=1, value=label)
    c.font = _font(bold=True, color=CLR_WHITE, size=10)
    c.fill = _fill(CLR_SECTION); c.alignment = _align("left"); c.border = _border()
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def score_band(s):
    if s <= 3:   return "Weak (0–3)"
    elif s <= 6: return "Moderate (4–6)"
    else:        return "Strong (7–10)"

def cohen_d(x):
    m = np.mean(x); s = np.std(x, ddof=1)
    return round(m / s, 4) if s > 0 else np.nan

def permutation_test(rm, rf, n=N_PERMS):
    obs = np.mean(rm) - np.mean(rf)
    rng = np.random.default_rng(42)
    count = 0
    for _ in range(n):
        swap = rng.binomial(1, 0.5, size=len(rm)).astype(bool)
        pm = np.where(swap, rf, rm); pf = np.where(swap, rm, rf)
        if abs(np.mean(pm) - np.mean(pf)) >= abs(obs):
            count += 1
    return round(obs, 4), round(count / n, 6)

def ras(row):
    s = pd.Series({"male": row.score_male, "female": row.score_female, "neutral": row.score_neutral})
    r = s.rank(ascending=False, method="average")
    return pd.Series({"rank_male": r["male"], "rank_female": r["female"], "rank_neutral": r["neutral"]})

def severity(rg):
    ag = abs(rg)
    if ag == 2.0:   return "Most Biased"
    elif ag == 1.5: return "Clearly Biased"
    elif ag == 1.0: return "Mildly Biased"
    else:           return "No Bias"

def build_pivot(df_slice):
    has_domain = "domain" in df_slice.columns
    pivot = df_slice.pivot_table(
        index=["cv_id", "technique"] + (["domain"] if has_domain else []),
        columns="version", values="score", aggfunc="mean"
    ).reset_index()
    pivot.columns.name = None
    pivot = pivot.rename(columns={"original": "score_neutral", "male": "score_male", "female": "score_female"})
    pivot = pivot.dropna(subset=["score_male", "score_female", "score_neutral"])
    pivot["delta"] = pivot["score_male"] - pivot["score_female"]
    pivot = pd.concat([pivot, pivot.apply(ras, axis=1)], axis=1)
    pivot["rank_gap"] = pivot["rank_male"] - pivot["rank_female"]
    pivot["severity"] = pivot["rank_gap"].apply(severity)
    pivot["score_band"] = pivot["score_neutral"].apply(score_band)
    pivot["abs_delta"] = pivot["delta"].abs()
    return pivot

def analyse_techniques(pivot):
    results = {}
    for tech in TECHNIQUES:
        t = pivot[pivot["technique"] == tech].copy()
        n = len(t)
        if n == 0:
            continue
        d = t["delta"]
        mean_m = round(t.score_male.mean(), 4); mean_f = round(t.score_female.mean(), 4)
        mean_n = round(t.score_neutral.mean(), 4); mean_d = round(d.mean(), 4)
        std_d  = round(d.std(ddof=1), 4)
        pct_m  = round((d > 0).sum() / n * 100, 2); pct_f = round((d < 0).sum() / n * 100, 2)
        pct_eq = round((d == 0).sum() / n * 100, 2)
        avg_rm = round(t.rank_male.mean(), 4); avg_rf = round(t.rank_female.mean(), 4)
        avg_rn = round(t.rank_neutral.mean(), 4); rg = round(t.rank_gap.mean(), 4)
        sev    = (t.severity.value_counts() / n * 100).round(2).to_dict()
        dci    = round(abs(pct_f - pct_m) / 100, 4)
        obs_rg, perm_p = permutation_test(t.rank_male.values, t.rank_female.values)
        parity_gap = round(abs(mean_m - mean_f), 4)
        results[tech] = {
            "n": n, "mean_m": mean_m, "mean_f": mean_f, "mean_n": mean_n,
            "mean_d": mean_d, "std_d": std_d,
            "pct_m": pct_m, "pct_f": pct_f, "pct_eq": pct_eq,
            "avg_rm": avg_rm, "avg_rf": avg_rf, "avg_rn": avg_rn, "rg": rg,
            "severity": sev, "dci": dci,
            "obs_rg": obs_rg, "perm_p": perm_p, "parity_gap": parity_gap,
        }
    return results

def analyse_domains(pivot):
    DOMAINS = sorted(pivot["domain"].dropna().unique()) if "domain" in pivot.columns else []
    domain_results = {}
    for domain in DOMAINS:
        for tech in TECHNIQUES:
            t = pivot[(pivot["domain"] == domain) & (pivot["technique"] == tech)].copy()
            n = len(t)
            if n == 0: continue
            d = t["delta"]
            domain_results[(domain, tech)] = {
                "n": n,
                "mean_m": round(t.score_male.mean(), 4), "mean_f": round(t.score_female.mean(), 4),
                "mean_n": round(t.score_neutral.mean(), 4), "mean_d": round(d.mean(), 4),
                "std_d": round(d.std(ddof=1), 4) if n > 1 else np.nan,
                "pct_m": round((d > 0).sum() / n * 100, 2), "pct_f": round((d < 0).sum() / n * 100, 2),
                "pct_eq": round((d == 0).sum() / n * 100, 2),
                "rank_gap": round(t.rank_gap.mean(), 4),
                "parity_gap": round(abs(t.score_male.mean() - t.score_female.mean()), 4),
                "cohen_d": cohen_d(d.values),
            }
    return domain_results

def analyse_score_range(pivot):
    score_range_results = {}
    for tech in TECHNIQUES:
        t = pivot[pivot["technique"] == tech].copy()
        band_data = {}; groups = []
        for band in BANDS:
            b = t[t["score_band"] == band]; n = len(b); d = b["delta"]
            band_data[band] = {
                "n": n,
                "mean_d": round(d.mean(), 4) if n > 0 else np.nan,
                "std_d": round(d.std(ddof=1), 4) if n > 1 else np.nan,
                "pct_m": round((d > 0).sum() / n * 100, 2) if n > 0 else np.nan,
                "pct_f": round((d < 0).sum() / n * 100, 2) if n > 0 else np.nan,
                "rank_gap": round(b.rank_gap.mean(), 4) if n > 0 else np.nan,
                "parity_gap": round(abs(b.score_male.mean() - b.score_female.mean()), 4) if n > 0 else np.nan,
                "cohen_d": cohen_d(d.values) if n > 1 else np.nan,
            }
            if n > 0: groups.append(b["abs_delta"].values)
        if len(groups) >= 2 and all(len(g) > 0 for g in groups):
            kw_stat, kw_p = stats.kruskal(*groups)
            kw_stat = round(kw_stat, 4); kw_p = round(kw_p, 6)
        else:
            kw_stat = kw_p = np.nan
        mw = {}
        band_arrays = {b: t[t["score_band"] == b]["abs_delta"].values for b in BANDS}
        pairs = [("Weak (0–3)", "Moderate (4–6)"), ("Moderate (4–6)", "Strong (7–10)"), ("Weak (0–3)", "Strong (7–10)")]
        for b1, b2 in pairs:
            a1, a2 = band_arrays[b1], band_arrays[b2]
            if len(a1) > 0 and len(a2) > 0:
                u_stat, u_p = stats.mannwhitneyu(a1, a2, alternative="two-sided")
                mw[f"{b1} vs {b2}"] = (round(u_stat, 4), round(u_p, 6))
            else:
                mw[f"{b1} vs {b2}"] = (np.nan, np.nan)
        score_range_results[tech] = {"bands": band_data, "kw_stat": kw_stat, "kw_p": kw_p, "mw": mw}
    return score_range_results

df_full = pd.read_csv(CSV_PATH)
df_full = df_full[df_full["success"] == True]
df_full = df_full[df_full["version"].isin(["male", "female", "original"])]

JD_INDEXS = sorted(df_full["jd_index"].unique())
print(f"Found {len(JD_INDEXS)} JDs: {JD_INDEXS}")

all_results        = {}
all_domain_results = {}
all_sr_results     = {}

for jd_index in JD_INDEXS:
    print(f"  Analysing JD {jd_index}...")
    df_run = df_full[df_full["jd_index"] == jd_index].copy()
    pivot  = build_pivot(df_run)
    all_results[jd_index]        = analyse_techniques(pivot)
    all_domain_results[jd_index] = analyse_domains(pivot)
    all_sr_results[jd_index]     = analyse_score_range(pivot)

wb  = Workbook()
wb.remove(wb.active)

ws_sum = wb.create_sheet(title="Summary", index=0)
ws_sum.merge_cells("A1:L1")
c = ws_sum["A1"]
c.value = "JD Sensitivity — Summary (all JDs × techniques)"
c.font = _font(bold=True, color=CLR_WHITE, size=12); c.fill = _fill(CLR_HEADER); c.alignment = _align("center")

sum_headers = ["JD", "Technique", "n", "Mean Δ", "Std Δ",
               "M>F (%)", "F>M (%)", "Equal (%)",
               "Rank Gap", "Parity Gap", "Perm p-value", "Significant", "PRI (rank gap)"]
_header_row(ws_sum, 2, sum_headers, bg=CLR_SECTION)

row = 3
for jd_index in JD_INDEXS:
    res  = all_results[jd_index]
    base = res.get(BASELINE, {})
    for tech in TECHNIQUES:
        r = res.get(tech)
        if r is None: continue
        sig  = r['perm_p'] < BONF_THRESHOLD
        pri  = round(r['rg'] - base.get('rg', 0), 4) if tech != BASELINE else None
        bg   = CLR_GREEN if sig else (CLR_ALT if row % 2 == 0 else CLR_WHITE)
        _data_row(ws_sum, row, [
            f"JD {jd_index}", f"{tech}{'  [baseline]' if tech == BASELINE else ''}",
            r['n'], r['mean_d'], r['std_d'],
            r['pct_m'], r['pct_f'], r['pct_eq'],
            r['rg'], r['parity_gap'], r['perm_p'],
            "YES" if sig else "no",
            f"{pri:+.4f}" if pri is not None else "—",
        ], bg=bg)
        row += 1
    row += 1

note_row = row + 1
ws_sum.merge_cells(f"A{note_row}:M{note_row}")
note = ws_sum.cell(row=note_row, column=1,
    value=f"Bonferroni threshold: α = {BONF_THRESHOLD} ({len(TECHNIQUES)} techniques).  Green = significant after correction.")
note.font = _font(color="595959", size=9); note.alignment = _align("left", wrap=True)
_set_col_widths(ws_sum, [10, 22, 6, 10, 10, 9, 9, 9, 10, 12, 14, 12, 14])

METRIC_ROWS = [
    ("n",                          "n"),
    ("Mean score — male",          "mean_m"),
    ("Mean score — female",        "mean_f"),
    ("Mean score — neutral",       "mean_n"),
    ("Mean Δ (male − female)",     "mean_d"),
    ("Std Δ",                      "std_d"),
    ("Male > female (%)",          "pct_m"),
    ("Female > male (%)",          "pct_f"),
    ("Equal (%)",                  "pct_eq"),
    ("Avg rank — male",            "avg_rm"),
    ("Avg rank — female",          "avg_rf"),
    ("Avg rank — neutral",         "avg_rn"),
    ("Rank gap (male − female)",   "rg"),
    ("Most biased (%)",            "sev_most"),
    ("Clearly biased (%)",         "sev_clearly"),
    ("Mildly biased (%)",          "sev_mildly"),
    ("No bias (%)",                "sev_none"),
    ("DCI",                        "dci"),
    ("Observed rank gap (perm)",   "obs_rg"),
    ("Perm p-value",               "perm_p"),
    (f"Significant (α={BONF_THRESHOLD})", "sig"),
    ("Parity gap",                 "parity_gap"),
    ("PRI — rank gap",             "pri_rg"),
    ("PRI — parity gap",           "pri_pg"),
]

for tech in TECHNIQUES:
    ws = wb.create_sheet(title=tech[:31])
    is_base = tech == BASELINE

    n_cols = 1 + len(JD_INDEXS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=f"{tech.upper()}{'  [BASELINE]' if is_base else ''}  —  metrics across JDs")
    c.font = _font(bold=True, color=CLR_WHITE, size=12); c.fill = _fill(CLR_HEADER); c.alignment = _align("center")

    _header_row(ws, 2, ["Metric"] + [f"JD {r}" for r in JD_INDEXS], bg=CLR_SECTION)

    def get_val(jd_index, key):
        r    = all_results[jd_index].get(tech, {})
        base = all_results[jd_index].get(BASELINE, {})
        if key == "sev_most":    return r.get("severity", {}).get("Most Biased", 0.0)
        if key == "sev_clearly": return r.get("severity", {}).get("Clearly Biased", 0.0)
        if key == "sev_mildly":  return r.get("severity", {}).get("Mildly Biased", 0.0)
        if key == "sev_none":    return r.get("severity", {}).get("No Bias", 0.0)
        if key == "sig":         return "YES" if r.get("perm_p", 1) < BONF_THRESHOLD else "no"
        if key == "pri_rg":
            return f"{round(r.get('rg',0) - base.get('rg',0), 4):+.4f}" if not is_base else "—"
        if key == "pri_pg":
            return f"{round(r.get('parity_gap',0) - base.get('parity_gap',0), 4):+.4f}" if not is_base else "—"
        return r.get(key, "")

    for i, (label, key) in enumerate(METRIC_ROWS, 3):
        bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
        values = [label] + [get_val(jd_index, key) for jd_index in JD_INDEXS]
        _data_row(ws, i, values, bg=bg)
        if key == "perm_p":
            for col_idx, jd_index in enumerate(JD_INDEXS, 2):
                val = all_results[jd_index].get(tech, {}).get("perm_p", 1)
                if val < BONF_THRESHOLD:
                    ws.cell(row=i, column=col_idx).fill = _fill(CLR_YELLOW)

    _set_col_widths(ws, [36] + [14] * len(JD_INDEXS))

ws_dom = wb.create_sheet(title="Domain Analysis")
row = 1
ws_dom.merge_cells(f"A{row}:N{row}")
c = ws_dom.cell(row=row, column=1,
    value="Domain Analysis — descriptive metrics + Cohen's d per JD × domain × technique")
c.font = _font(bold=True, color=CLR_WHITE, size=11); c.fill = _fill(CLR_HEADER); c.alignment = _align("center")
row += 1

dom_headers = ["JD", "Domain", "Technique", "n",
               "Mean score M", "Mean score F", "Mean score N",
               "Mean Δ", "Std Δ", "M>F (%)", "F>M (%)", "Equal (%)",
               "Rank gap", "Parity gap", "Cohen's d"]
_header_row(ws_dom, row, dom_headers, bg=CLR_SECTION)
row += 1

i = 0
for jd_index in JD_INDEXS:
    dr = all_domain_results[jd_index]
    prev_domain = None
    for (domain, tech), r in sorted(dr.items()):
        bg = CLR_ALT if i % 2 == 0 else CLR_WHITE
        domain_label = domain if domain != prev_domain else ""
        prev_domain  = domain
        _data_row(ws_dom, row, [
            f"JD {jd_index}", domain_label, tech, r["n"],
            r["mean_m"], r["mean_f"], r["mean_n"],
            r["mean_d"], r["std_d"], r["pct_m"], r["pct_f"], r["pct_eq"],
            r["rank_gap"], r["parity_gap"], r["cohen_d"],
        ], bg=bg)
        row += 1; i += 1
    row += 1

note_row = row
ws_dom.merge_cells(f"A{note_row}:O{note_row}")
note = ws_dom.cell(row=note_row, column=1,
    value="Cohen's d: |d|<0.2 negligible, 0.2–0.5 small, 0.5–0.8 medium, >0.8 large.  "
          "Formal significance testing omitted at domain level.")
note.font = _font(color="595959", size=9); note.alignment = _align("left", wrap=True)
ws_dom.row_dimensions[note_row].height = 28
_set_col_widths(ws_dom, [10, 18, 18, 5, 13, 13, 13, 10, 10, 9, 9, 9, 10, 12, 10])

ws_sr = wb.create_sheet(title="Score Range Analysis")
row = 1
ws_sr.merge_cells(f"A{row}:K{row}")
c = ws_sr.cell(row=row, column=1,
    value="Score Range Analysis — bias concentration by score band per JD × technique")
c.font = _font(bold=True, color=CLR_WHITE, size=11); c.fill = _fill(CLR_HEADER); c.alignment = _align("center")
row += 2

for jd_index in JD_INDEXS:
    ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
    c = ws_sr.cell(row=row, column=1, value=f"JD {jd_index}")
    c.font = _font(bold=True, color=CLR_WHITE, size=11); c.fill = _fill(CLR_HEADER); c.alignment = _align("left")
    row += 1

    for tech in TECHNIQUES:
        sr = all_sr_results[jd_index].get(tech, {})
        if not sr: continue

        ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)
        c = ws_sr.cell(row=row, column=1, value=f"{tech.upper()}")
        c.font = _font(bold=True, color=CLR_WHITE, size=10); c.fill = _fill(CLR_SECTION); c.alignment = _align("left")
        row += 1

        band_headers = ["Score band", "n", "Mean Δ", "Std Δ",
                        "M>F (%)", "F>M (%)", "Rank gap", "Parity gap", "Cohen's d"]
        _header_row(ws_sr, row, band_headers, bg="4472C4", col_start=1)
        row += 1

        for j, band in enumerate(BANDS):
            b  = sr["bands"][band]; bg = CLR_ALT if j % 2 == 0 else CLR_WHITE
            _data_row(ws_sr, row, [
                band, b["n"], b["mean_d"], b["std_d"],
                b["pct_m"], b["pct_f"], b["rank_gap"], b["parity_gap"], b["cohen_d"],
            ], bg=bg)
            row += 1

        kw_sig = (not np.isnan(sr["kw_p"])) and sr["kw_p"] < 0.05
        ws_sr.cell(row=row, column=1,
            value=f"Kruskal-Wallis:  H = {sr['kw_stat']}  |  p = {sr['kw_p']}"
                  f"  →  {'significant (p<0.05)' if kw_sig else 'not significant'}").font = _font(bold=kw_sig)
        if kw_sig: ws_sr.cell(row=row, column=1).fill = _fill(CLR_YELLOW)
        ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        row += 1

        for pair_label, (u_stat, u_p) in sr["mw"].items():
            mw_sig = (not np.isnan(u_p)) and u_p < 0.05
            ws_sr.cell(row=row, column=1,
                value=f"  Mann-Whitney {pair_label}:  U = {u_stat}  |  p = {u_p}"
                      f"  →  {'significant' if mw_sig else 'not significant'}").font = _font(color="595959", size=9)
            if mw_sig: ws_sr.cell(row=row, column=1).fill = _fill(CLR_YELLOW)
            ws_sr.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            row += 1
        row += 1
    row += 1

_set_col_widths(ws_sr, [22, 5, 10, 10, 9, 9, 10, 12, 10])

wb.save(OUTPUT_XLSX)
print(f"\nDone! Results saved to: {OUTPUT_XLSX}")
