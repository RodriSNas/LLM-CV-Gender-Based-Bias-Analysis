import re
import ast
import unicodedata
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

JD_CSV       = "JD_data.csv"
SAMPLED_XLSX = "C:/Users/rodri/Desktop/Mestrado/Tese/Method/datasets/Resume/selected_cvs_updated_expanded_v5.xlsx"
OUTPUT_XLSX  = "jd_selection_results_v5.xlsx"

TOP_N     = 10
MIN_CHARS = 600
MAX_CHARS = 2500

DOMAINS = [
    "FINANCE",
    "INFORMATION-TECHNOLOGY",
    "ENGINEERING",
    "HR",
    "TEACHER",
    "HEALTHCARE",
    "SALES",
    "DIGITAL-MEDIA"
]

# Maps each domain to the major_job categories in JD_data.csv
JD_DOMAIN_MAP = {
    "FINANCE":                ["BUSINESS AND ADMINISTRATION PROFESSIONALS"],
    "INFORMATION-TECHNOLOGY": ["INFORMATION AND COMMUNICATIONS TECHNOLOGY PROFESSIONALS"],
    "ENGINEERING":            ["SCIENCE AND ENGINEERING PROFESSIONALS"],
    "HR":                     ["BUSINESS AND ADMINISTRATION PROFESSIONALS"],
    "TEACHER":                ["TEACHING PROFESSIONALS"],
    "HEALTHCARE":             ["HEALTH ASSOCIATE PROFESSIONALS", "HEALTH PROFESSIONALS"],
    "SALES":                  ["BUSINESS AND ADMINISTRATION ASSOCIATE PROFESSIONALS", "BUSINESS AND ADMINISTRATION PROFESSIONALS"],
    "DIGITAL-MEDIA":          ["INFORMATION AND COMMUNICATIONS TECHNICIANS","BUSINESS AND ADMINISTRATION PROFESSIONALS"]
}

JD_TITLE_FILTER = {
    "INFORMATION-TECHNOLOGY": [
        "system engineer", "system designer", "system developer", "system programmer",
        "application developer", "application engineer", "database engineer",
        "database developer", "software designer", "security analyst",
        "data analyst", "data engineer",
    ],
    "SALES": [
        "sales manager", "account manager", "sales executive",
        "sales representative", "business development manager",
        "sales consultant", "field sales representative",
        "sales advisor", "sales director", "account executive",
    ],
    "ENGINEERING": [
        "mechanical engineer", "project engineer", "process engineer",
        "verification engineer", "development engineer", "maintenance engineer",
        "instrument engineer", "mechatronics engineer", "plant engineer",
        "testing engineer", "environmental engineer", "construction planner",
        "mechanical designer", "inspection engineer",
    ],
    "TEACHER": [
        "language teacher", "music teacher", "primary school teacher", "supply teacher",
        "geography teacher", "special needs teacher", "drama teacher",
        "art teacher", "computing teacher",
    ],
    "FINANCE": [
        "financial analyst", "auditor", "accountant",
    ],
    "HR": [
        "recruitment researcher", "recruitment officer",
    ],
    "DIGITAL-MEDIA": [
        "video editor", "copywriter", "media planner", "ux designer"
    ],
}

QUALITY_WORDS = [
    "experience", "qualification", "degree", "skills", "knowledge",
    "required", "essential", "ability", "proficiency", "background",
    "demonstrated", "proven", "responsibilities", "duties",
]

STOPWORDS = {
    "and","or","the","a","an","in","of","to","for","with","is","are","be",
    "will","must","can","able","have","has","you","your","we","our","their",
    "this","that","as","at","on","by","from","not","may","its","it","do",
    "been","all","which","who","what","when","where","how","they","them",
    "also","into","such","more","than","well",
}

def parse_description(raw: str) -> str:
    try:
        parsed = ast.literal_eval(raw)
        if isinstance(parsed, list):
            raw = " ".join(str(x) for x in parsed)
    except Exception:
        pass
    raw = str(raw).strip("[]").strip("'").strip()
    raw = re.sub(r"\s{3,}", "  ", raw)
    return raw


def normalise_for_dedup(text: str) -> str:
    """Normalise text for near-duplicate detection."""
    text = unicodedata.normalize("NFKC", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    text = re.sub(r"[^\w\s]", "", text)
    return text


def passes_quality_gate(text: str) -> bool:
    """True if the JD contains ≥2 candidate-requirement signal words."""
    t = text.lower()
    return sum(1 for w in QUALITY_WORDS if w in t) >= 2


def score_cv_vs_jd(cv_text: str, jd_text: str) -> float:
    jd_words = set(
        w for w in re.findall(r"\b[a-z]{4,}\b", jd_text.lower())
        if w not in STOPWORDS
    )
    cv_words = set(
        w for w in re.findall(r"\b[a-z]{4,}\b", cv_text.lower())
        if w not in STOPWORDS
    )
    if not jd_words:
        return 0.0
    overlap = len(jd_words & cv_words) / len(jd_words)
    return round(min(overlap * 25, 10), 1)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def _header_row(ws, row: int, values: list, bg="1F3864", fg="FFFFFF") -> None:
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = _font(bold=True, color=fg)
        c.fill      = _fill(bg)
        c.alignment = _align("center")
        c.border    = _border()


def load_jds(csv_path: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path)
    df["description"] = df["description"].apply(parse_description)
    df["jd_length"]   = df["description"].str.len()
    return df


def load_sampled_cvs(xlsx_path: str) -> dict[str, list[dict]]:
    """Returns {domain: [{"id": ..., "text": ...}, ...]}"""
    wb  = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws  = wb["Neutral CVs"]
    rows = list(ws.iter_rows(values_only=True))
    headers  = rows[0]
    cat_idx  = headers.index("Category")
    text_idx = headers.index("Resume Text")
    id_idx   = headers.index("ID")

    cvs: dict[str, list] = {}
    for row in rows[1:]:
        domain = row[cat_idx]
        if domain not in cvs:
            cvs[domain] = []
        cvs[domain].append({"id": row[id_idx], "text": str(row[text_idx])})
    wb.close()
    return cvs


def select_top_jds(
    jd_df: pd.DataFrame,
    domain: str,
    sampled_cvs: list[dict],
) -> list[dict]:
    #Category filter
    cats = JD_DOMAIN_MAP[domain]
    sub  = jd_df[jd_df["major_job"].isin(cats)].copy()

    #Job-title filter
    titles = JD_TITLE_FILTER.get(domain)
    if titles:
        sub = sub[sub["job"].isin(titles)]

    #Length band
    sub = sub[(sub["jd_length"] >= MIN_CHARS) & (sub["jd_length"] <= MAX_CHARS)]

    #Quality gate
    sub = sub[sub["description"].apply(passes_quality_gate)]

    if sub.empty:
        print(f"  ⚠  {domain}: no JDs passed all filters.")
        return []

    #Deduplicate by normalised text
    sub = sub.copy()
    sub["_norm"] = sub["description"].apply(normalise_for_dedup)
    sub = sub.drop_duplicates(subset="_norm", keep="first").drop(columns="_norm")

    available = len(sub)
    if available < TOP_N:
        print(f"  ⚠  {domain}: only {available} unique JDs available "
              f"(wanted {TOP_N}) — consider relaxing title list or MAX_CHARS.")

    #Score each JD against every sampled CV, rank by mean keyword overlap
    results = []
    for _, jd_row in sub.iterrows():
        jd_text = jd_row["description"]
        scores  = [score_cv_vs_jd(cv["text"], jd_text) for cv in sampled_cvs]

        results.append({
            "position":      jd_row.get("position", ""),
            "job":           jd_row["job"],
            "jd_length":     jd_row["jd_length"],
            "mean_overlap":  round(sum(scores) / len(scores), 4),
            "jd_text":       jd_text,
        })

    results.sort(key=lambda x: x["mean_overlap"], reverse=True)
    return results[:TOP_N]


def export_excel(all_results: dict[str, list[dict]], sampled_cvs: dict) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    #Summary sheet
    ws_sum = wb.create_sheet(title="Summary")
    ws_sum.merge_cells("A1:I1")
    c = ws_sum["A1"]
    c.value     = (f"JD Selection Results — top {TOP_N} per domain, "
                   f"scored against sampled CVs  |  length {MIN_CHARS}–{MAX_CHARS} chars")
    c.font      = _font(bold=True, color="FFFFFF", size=12)
    c.fill      = _fill("1F3864")
    c.alignment = _align("center")

    _header_row(ws_sum, 2,
        ["Domain", "Rank", "Position", "Job Type", "Mean Overlap", "Length"],
        bg="2E75B6")

    row = 3
    for domain in DOMAINS:
        results = all_results.get(domain, [])
        if not results:
            continue
        for rank, r in enumerate(results, 1):
            bg = "EAF3DE" if rank == 1 else ("D9E2F3" if row % 2 == 0 else "FFFFFF")
            vals = [
                domain if rank == 1 else "", rank,
                r["position"][:60], r["job"],
                r["mean_overlap"], r["jd_length"],
            ]
            for col, v in enumerate(vals, 1):
                c = ws_sum.cell(row=row, column=col, value=v)
                c.font      = _font(bold=(rank == 1))
                c.fill      = _fill(bg)
                c.alignment = _align(
                    "center" if col not in (3,) else "left"
                )
                c.border    = _border()
            row += 1
        row += 1  # blank spacer between domains

    note_text = (
        "Green rows = top-ranked JD per domain.  "
        "Mean Overlap: average keyword overlap score (0–10) between each JD and all sampled CVs in the domain.  "
        f"Ranked highest-to-lowest overlap — top 10 selected per domain.  "
        f"Filters: job-title allowlist + {MIN_CHARS}–{MAX_CHARS} chars + ≥2 requirement words."
    )
    note = ws_sum.cell(row=row, column=1, value=note_text)
    note.font      = _font(color="595959", size=9)
    note.alignment = _align("left", wrap=True)
    ws_sum.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=9)
    ws_sum.row_dimensions[row].height = 30

    for col, w in enumerate([22, 6, 42, 20, 14, 8], 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    #Per-domain detail sheets
    for domain in DOMAINS:
        results = all_results.get(domain, [])
        if not results:
            continue

        n_cvs = len(sampled_cvs.get(domain, []))
        ws    = wb.create_sheet(title=domain[:31])

        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value     = (f"{domain} — top {len(results)} JDs  |  "
                       f"scored against {n_cvs} sampled CVs")
        c.font      = _font(bold=True, color="FFFFFF", size=11)
        c.fill      = _fill("1F3864")
        c.alignment = _align("center")

        row = 2
        for rank, r in enumerate(results, 1):
            ws.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=3
            )
            c = ws.cell(
                row=row, column=1,
                value=f"#{rank}  [overlap={r['mean_overlap']:.4f}]  "
                      f"{r['position']}  ({r['job']}  |  {r['jd_length']} chars)"
            )
            c.font      = _font(bold=True, color="FFFFFF")
            c.fill      = _fill("3B6D11" if rank == 1 else "2E75B6")
            c.alignment = _align("left")
            row += 1

            metrics = [
                ("Mean keyword overlap (0–10)", r["mean_overlap"]),
                ("JD length (chars)",           r["jd_length"]),
            ]
            for i, (label, val) in enumerate(metrics):
                bg = "F2F2F2" if i % 2 == 0 else "FFFFFF"
                for col, v in enumerate([label, val], 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font      = _font()
                    c.fill      = _fill(bg)
                    c.border    = _border()
                    c.alignment = _align()
                row += 1

            c_label = ws.cell(row=row, column=1, value="JD text")
            c_label.font      = _font(bold=True)
            c_label.fill      = _fill("F2F2F2")
            c_label.border    = _border()
            c_label.alignment = _align()

            c_text = ws.cell(row=row, column=2, value=r["jd_text"])
            c_text.font      = _font()
            c_text.fill      = _fill("FFFFFF")
            c_text.border    = _border()
            c_text.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )
            ws.merge_cells(
                start_row=row, start_column=2, end_row=row, end_column=3
            )
            ws.row_dimensions[row].height = 150
            row += 2

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 65
        ws.column_dimensions["C"].width = 20

    wb.save(OUTPUT_XLSX)


def main() -> None:
    print("=" * 60)
    print("JD SELECTION — FINAL")
    print("=" * 60)

    print(f"\n[1/4] Loading JDs from {JD_CSV}...")
    jd_df = load_jds(JD_CSV)
    print(f"      {len(jd_df)} JDs loaded")

    print(f"\n[2/4] Loading sampled CVs from {SAMPLED_XLSX}...")
    sampled_cvs = load_sampled_cvs(SAMPLED_XLSX)
    for d, cvs in sampled_cvs.items():
        print(f"      {d}: {len(cvs)} CVs")

    print(f"\n[3/4] Filtering and scoring JDs...")
    print(f"      Length: {MIN_CHARS}–{MAX_CHARS} chars | Quality gate: ≥2 requirement words")
    all_results: dict[str, list[dict]] = {}

    for domain in DOMAINS:
        cvs = sampled_cvs.get(domain, [])
        if not cvs:
            print(f"  ⚠  {domain}: no sampled CVs found — skipping")
            continue

        results = select_top_jds(jd_df, domain, cvs)
        all_results[domain] = results

        if results:
            top = results[0]
            print(
                f"      {domain}: {len(results)} JDs selected | "
                f"top → [{top['mean_overlap']:.4f}] {top['job']} "
                f"({top['jd_length']} chars)"
            )

    print(f"\n[4/4] Exporting to {OUTPUT_XLSX}...")
    export_excel(all_results, sampled_cvs)
    print(f"      Done — {OUTPUT_XLSX} saved.")
    print("\n" + "=" * 60)


if __name__ == "__main__":
    main()
